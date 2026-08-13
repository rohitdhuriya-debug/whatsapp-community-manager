"""Turn model-generated JSON into a real PDF or Excel file.

The model decides the structure (sections/bullets for a document, columns/rows
for a sheet); this module owns the typography and layout so every asset looks
like it came from the same brand rather than from a language model.
"""

from __future__ import annotations

import base64
import logging
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    ListFlowable,
    ListItem,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from ..config import BASE_DIR, config

log = logging.getLogger(__name__)

ASSET_DIR = BASE_DIR / "assets"

# Matches the dashboard palette (HC-4) so assets look like the same product.
INK = colors.HexColor("#111827")
MUTED = colors.HexColor("#6B7280")
ACCENT = colors.HexColor("#4F46E5")
LINE = colors.HexColor("#E5E7EB")
WASH = colors.HexColor("#F9FAFB")

PDF_MIME = "application/pdf"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_FONTS_READY = False
FONT_REGULAR = "Helvetica"
FONT_BOLD = "Helvetica-Bold"
FONT_UNICODE = "Helvetica"
FONT_DISPLAY = "Helvetica-Bold"

MAC_FONTS = {
    "AppSans": "/System/Library/Fonts/Supplemental/Arial.ttf",
    "AppSans-Bold": "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
    # Covers Devanagari and most other scripts, for non-English assets.
    "AppUnicode": "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    # Heavier face for cover headlines. Narrow coverage (Latin only), so it is
    # only ever used for display text that has been through _fit_to_font.
    "AppDisplay": "/System/Library/Fonts/Supplemental/Arial Black.ttf",
}


def _register_fonts() -> None:
    """Register TTFs once. Falls back to built-in Helvetica if absent."""
    global _FONTS_READY, FONT_REGULAR, FONT_BOLD, FONT_UNICODE, FONT_DISPLAY
    if _FONTS_READY:
        return

    for name, path in MAC_FONTS.items():
        if not Path(path).exists():
            log.warning("Font missing, falling back to Helvetica: %s", path)
            continue
        try:
            pdfmetrics.registerFont(TTFont(name, path))
        except Exception as exc:
            log.warning("Could not register %s: %s", name, exc)

    registered = set(pdfmetrics.getRegisteredFontNames())
    if {"AppSans", "AppSans-Bold"} <= registered:
        FONT_REGULAR, FONT_BOLD = "AppSans", "AppSans-Bold"
    if "AppUnicode" in registered:
        FONT_UNICODE = "AppUnicode"
    if "AppDisplay" in registered:
        FONT_DISPLAY = "AppDisplay"
    else:
        FONT_DISPLAY = FONT_BOLD

    # Last, not first: setting the flag before the globals are assigned lets a
    # concurrent caller return early and draw everything in Helvetica.
    _FONTS_READY = True


def _needs_unicode_font(text: str) -> bool:
    """True when the text leaves Latin-1 - Devanagari, CJK, and friends.

    reportlab has no per-glyph fallback, so a document renders entirely in
    Arial (crisp, has a real bold) or entirely in Arial Unicode (wide script
    coverage, no bold face). Pick once, per document.
    """
    for ch in text:
        if ord(ch) < 0x0100:
            continue
        if unicodedata.category(ch).startswith(("L", "N")):
            return True
    return False


# Characters no font available here can draw, mapped to something readable.
# The rupee sign is the big one: it was added to Unicode in 2010, so neither
# Arial nor Arial Unicode MS has it, and it would silently render as a box.
GLYPH_FALLBACKS = {
    "₹": "Rs.",   # ₹
    "–": "-",     # – en dash
    "—": "-",     # — em dash
    "‘": "'", "’": "'",
    "“": '"', "”": '"',
    "•": "-",     # • bullet
    "…": "...",
    " ": " ",
    "​": "", "️": "",
    "→": "->", "←": "<-",
    "✅": "", "❌": "",
}


def _coverage(font_name: str) -> set[int]:
    """Codepoints the registered font can actually draw."""
    try:
        face = pdfmetrics.getFont(font_name).face
    except Exception:
        return set()
    mapping = getattr(face, "charToGlyph", None)
    return set(mapping.keys()) if isinstance(mapping, dict) else set()


def _fit_to_font(text: str, font_name: str) -> str:
    """Replace anything the font cannot draw, so a box never reaches the page.

    Built-in Helvetica has no charToGlyph table; in that case fall back to the
    static map plus an ASCII-only filter.
    """
    if not text:
        return ""
    covered = _coverage(font_name)

    out: list[str] = []
    for ch in text:
        code = ord(ch)
        if covered and code in covered:
            out.append(ch)
            continue
        if ch in GLYPH_FALLBACKS:
            out.append(GLYPH_FALLBACKS[ch])
            continue
        if not covered:
            # Unknown coverage: keep Latin-1, drop the rest.
            out.append(ch if code < 0x0100 else "")
            continue
        # Last resort: strip accents, else drop (emoji, rare symbols).
        decomposed = unicodedata.normalize("NFKD", ch)
        ascii_form = "".join(c for c in decomposed if ord(c) < 0x80)
        out.append(ascii_form)

    return re.sub(r"[ \t]{2,}", " ", "".join(out)).strip()


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def slugify(text: str, fallback: str = "asset") -> str:
    text = unicodedata.normalize("NFKD", text or "")
    text = re.sub(r"[^\w\s-]", "", text).strip().lower()
    text = re.sub(r"[\s_-]+", "-", text)
    return text[:60].strip("-") or fallback


def _clean(value: Any) -> str:
    """Model output is prose - strip stray markdown so it does not render raw."""
    text = "" if value is None else str(value)
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)
    text = re.sub(r"(?<!\w)[*_]{1,2}(?=\w)", "", text)
    text = re.sub(r"^\s{0,3}#{1,6}\s+", "", text, flags=re.MULTILINE)
    return " ".join(text.split())


def _escape(text: str) -> str:
    """Paragraph() parses a mini-HTML, so raw & < > must be escaped."""
    return (
        text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    )


def _timestamp() -> str:
    return datetime.now(config.tz).strftime("%d %b %Y")


def file_to_b64(path: Path) -> str:
    return base64.b64encode(path.read_bytes()).decode()


def human_size(num_bytes: int) -> str:
    for unit in ("B", "KB", "MB"):
        if num_bytes < 1024 or unit == "MB":
            return f"{num_bytes:.0f} {unit}" if unit == "B" else f"{num_bytes:.1f} {unit}"
        num_bytes /= 1024.0
    return f"{num_bytes:.1f} MB"


# ---------------------------------------------------------------------------
# Previews - so a generated file can be inspected without downloading it
# ---------------------------------------------------------------------------


def preview(path_str: str, kind: str, max_rows: int = 40) -> dict[str, Any]:
    """Structured summary of a generated file, for rendering in the dashboard."""
    path = Path(path_str)
    if not path.exists():
        return {"kind": kind, "error": "File is no longer on disk."}
    try:
        if kind == "excel":
            return _preview_excel(path, max_rows)
        return _preview_pdf(path)
    except Exception as exc:  # a preview failure must never break generation
        log.warning("Could not preview %s: %s", path.name, exc)
        return {"kind": kind, "error": f"Preview unavailable ({exc})."}


def _preview_excel(path: Path, max_rows: int) -> dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in workbook.worksheets:
        rows = [
            ["" if c is None else str(c) for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
        rows = [r for r in rows if any(str(c).strip() for c in r)]

        # Row 1 is the title and row 3 the header (see build_excel).
        title = rows[0][0] if rows else ""
        header, body = [], []
        for index, row in enumerate(rows):
            if index == 0:
                continue
            if not header and sum(1 for c in row if str(c).strip()) > 1:
                header = row
            elif header:
                body.append(row)

        width = len(header) or max((len(r) for r in body), default=0)
        sheets.append({
            "name": ws.title,
            "title": title,
            "columns": header[:width],
            "rows": [r[:width] for r in body[:max_rows]],
            "total_rows": len(body),
            "truncated": len(body) > max_rows,
        })
    workbook.close()
    return {"kind": "excel", "sheets": sheets,
            "size": human_size(path.stat().st_size)}


def render_page(path: Path, index: int, dpi: int = 110) -> bytes:
    """Rasterise one PDF page to PNG.

    Embedding a PDF in an <iframe> depends on the browser's PDF plugin, which
    is not always available (and renders blank when it is not). Serving page
    images instead makes the preview work everywhere.
    """
    import pymupdf

    with pymupdf.open(str(path)) as doc:
        if index < 0 or index >= doc.page_count:
            raise IndexError(f"Page {index + 1} is out of range (1-{doc.page_count}).")
        return doc[index].get_pixmap(dpi=dpi).tobytes("png")


def cover_image(path: Path, dpi: int = 130) -> tuple[bytes, str] | None:
    """First page of a PDF as a PNG, for use as a post preview image.

    A document attachment shows only a filename in the feed; leading with a
    picture of page one is what makes people actually open it.
    """
    if path.suffix.lower() != ".pdf":
        return None
    try:
        png = render_page(path, 0, dpi=dpi)
    except Exception as exc:
        log.warning("Could not build cover image for %s: %s", path.name, exc)
        return None
    return png, f"{path.stem}-preview.png"


def page_count(path: Path) -> int:
    try:
        import pymupdf

        with pymupdf.open(str(path)) as doc:
            return doc.page_count
    except Exception:
        return 0


def _preview_pdf(path: Path) -> dict[str, Any]:
    pages = None
    outline: list[str] = []
    try:
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        pages = len(reader.pages)
        text = "\n".join((p.extract_text() or "") for p in reader.pages[:3])
        # Heading-ish lines: short, title-cased, not sentences.
        for line in (l.strip() for l in text.splitlines()):
            if 3 < len(line) < 70 and not line.endswith((".", ",")) and line[:1].isupper():
                if line not in outline:
                    outline.append(line)
            if len(outline) >= 12:
                break
    except Exception as exc:
        log.debug("PDF outline unavailable: %s", exc)

    return {"kind": "pdf", "pages": pages, "outline": outline,
            "size": human_size(path.stat().st_size)}


def list_assets(limit: int = 60) -> list[dict[str, Any]]:
    """Everything generated so far, newest first."""
    ASSET_DIR.mkdir(parents=True, exist_ok=True)
    files = [p for p in ASSET_DIR.iterdir() if p.is_file() and not p.name.startswith(".")]
    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)

    out = []
    for path in files[:limit]:
        stat = path.stat()
        out.append({
            "filename": path.name,
            "kind": "excel" if path.suffix in (".xlsx", ".xls") else
                    "pdf" if path.suffix == ".pdf" else path.suffix.lstrip("."),
            "size": human_size(stat.st_size),
            "bytes": stat.st_size,
            "modified": datetime.fromtimestamp(stat.st_mtime, config.tz).isoformat(),
            "modified_label": datetime.fromtimestamp(stat.st_mtime, config.tz)
                              .strftime("%d %b %Y, %I:%M %p"),
        })
    return out


def resolve_asset(filename: str) -> Path:
    """Map a filename to a path inside ASSET_DIR, refusing traversal."""
    candidate = (ASSET_DIR / Path(filename).name).resolve()
    if candidate.parent != ASSET_DIR.resolve():
        raise ValueError("Invalid asset path.")
    if not candidate.exists():
        raise FileNotFoundError(filename)
    return candidate


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


class _DocTemplate(BaseDocTemplate):
    """Adds the footer rule and page numbers to every page."""

    def __init__(self, *args: Any, footer: str = "", **kwargs: Any):
        super().__init__(*args, **kwargs)
        self.footer_text = footer
        frame = Frame(
            22 * mm, 20 * mm, self.width, self.height,
            leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
            id="body",
        )
        self.addPageTemplates([PageTemplate(id="main", frames=[frame], onPage=self._decorate)])

    def _decorate(self, canvas: Any, doc: Any) -> None:
        canvas.saveState()
        canvas.setStrokeColor(LINE)
        canvas.setLineWidth(0.6)
        canvas.line(22 * mm, 16 * mm, A4[0] - 22 * mm, 16 * mm)

        canvas.setFont(FONT_REGULAR, 8)
        canvas.setFillColor(MUTED)
        if self.footer_text:
            canvas.drawString(22 * mm, 11 * mm, self.footer_text[:110])
        canvas.drawRightString(A4[0] - 22 * mm, 11 * mm, str(canvas.getPageNumber()))
        canvas.restoreState()


def _pdf_styles(unicode_mode: bool) -> dict[str, ParagraphStyle]:
    body_font = FONT_UNICODE if unicode_mode else FONT_REGULAR
    bold_font = FONT_UNICODE if unicode_mode else FONT_BOLD

    return {
        "title": ParagraphStyle(
            "title", fontName=bold_font, fontSize=25, leading=30,
            textColor=INK, spaceAfter=5,
        ),
        "subtitle": ParagraphStyle(
            "subtitle", fontName=body_font, fontSize=11.5, leading=17,
            textColor=MUTED, spaceAfter=3,
        ),
        "heading": ParagraphStyle(
            "heading", fontName=bold_font, fontSize=14, leading=19,
            textColor=INK, spaceBefore=17, spaceAfter=7,
        ),
        "body": ParagraphStyle(
            "body", fontName=body_font, fontSize=10.5, leading=16.5,
            textColor=INK, alignment=TA_LEFT, spaceAfter=7,
        ),
        "bullet": ParagraphStyle(
            "bullet", fontName=body_font, fontSize=10.5, leading=16,
            textColor=INK, spaceAfter=4,
        ),
        "kicker": ParagraphStyle(
            "kicker", fontName=bold_font, fontSize=8.5, leading=12,
            textColor=ACCENT, spaceAfter=8,
        ),
        "note": ParagraphStyle(
            "note", fontName=body_font, fontSize=9.5, leading=14.5,
            textColor=MUTED, spaceBefore=6,
        ),
    }


def build_pdf(payload: dict[str, Any], *, brand: str = "") -> tuple[Path, str]:
    """Render the document JSON to a PDF. Returns (path, filename)."""
    _register_fonts()
    ASSET_DIR.mkdir(parents=True, exist_ok=True)

    title = _clean(payload.get("title")) or "Resource"
    subtitle = _clean(payload.get("subtitle"))
    sections = payload.get("sections") or []
    note = _clean(payload.get("footer_note"))

    # Decide the font family once, from the whole document.
    corpus = " ".join(
        [title, subtitle, note]
        + [f"{s.get('heading','')} {s.get('body','')} {' '.join(s.get('bullets') or [])}"
           for s in sections if isinstance(s, dict)]
    )
    unicode_mode = _needs_unicode_font(corpus)
    styles = _pdf_styles(unicode_mode)
    body_font = FONT_UNICODE if unicode_mode else FONT_REGULAR

    def prep(value: Any) -> str:
        """Clean -> make renderable in the chosen font -> escape for Paragraph."""
        return _escape(_fit_to_font(_clean(value), body_font))

    filename = f"{slugify(title)}-{datetime.now(config.tz):%Y%m%d}.pdf"
    path = ASSET_DIR / filename

    story: list[Any] = []
    if brand:
        story.append(Paragraph(prep(brand.upper()), styles["kicker"]))
    story.append(Paragraph(prep(title), styles["title"]))
    if subtitle:
        story.append(Paragraph(prep(subtitle), styles["subtitle"]))
    story.append(Spacer(1, 4 * mm))
    story.append(
        Table([[""]], colWidths=[166 * mm], rowHeights=[1.1],
              style=TableStyle([("BACKGROUND", (0, 0), (-1, -1), ACCENT)]))
    )
    story.append(Spacer(1, 3 * mm))

    for section in sections:
        if not isinstance(section, dict):
            continue
        heading = prep(section.get("heading"))
        body = prep(section.get("body"))
        bullets = [b for b in (prep(x) for x in section.get("bullets") or []) if b]
        table = section.get("table")

        if heading:
            story.append(Paragraph(heading, styles["heading"]))
        if body:
            story.append(Paragraph(body, styles["body"]))
        if bullets:
            story.append(
                ListFlowable(
                    [ListItem(Paragraph(b, styles["bullet"]), leftIndent=12)
                     for b in bullets],
                    bulletType="bullet", bulletColor=ACCENT, bulletFontSize=7,
                    leftIndent=13, spaceAfter=6,
                )
            )
        if isinstance(table, dict) and table.get("columns"):
            story.append(_pdf_table(table, unicode_mode))
        if section.get("page_break"):
            story.append(PageBreak())

    if note:
        story.append(Spacer(1, 6 * mm))
        story.append(Paragraph(prep(note), styles["note"]))

    footer = _fit_to_font(f"{brand + ' · ' if brand else ''}{_timestamp()}", body_font)
    doc = _DocTemplate(
        str(path), pagesize=A4,
        leftMargin=22 * mm, rightMargin=22 * mm, topMargin=22 * mm, bottomMargin=22 * mm,
        title=title, author=brand or "Upsurge WhatsApp Manager", footer=footer,
    )
    # build() consumes the story list, so measure it first.
    flowables = len(story)
    doc.build(story)
    log.info("Built PDF %s from %s flowables", filename, flowables)
    return path, filename


def _pdf_table(table: dict[str, Any], unicode_mode: bool) -> Table:
    font = FONT_UNICODE if unicode_mode else FONT_REGULAR
    bold = FONT_UNICODE if unicode_mode else FONT_BOLD
    def prep(value: Any) -> str:
        return _escape(_fit_to_font(_clean(value), font))

    columns = [prep(c) for c in table.get("columns") or []]
    rows = [[prep(c) for c in row] for row in table.get("rows") or []]

    style = ParagraphStyle("cell", fontName=font, fontSize=9, leading=12.5, textColor=INK)
    head_style = ParagraphStyle("cellh", fontName=bold, fontSize=9, leading=12.5,
                                textColor=colors.white)

    data = [[Paragraph(c, head_style) for c in columns]]
    for row in rows:
        padded = (row + [""] * len(columns))[: len(columns)]
        data.append([Paragraph(c, style) for c in padded])

    width = 166 * mm / max(len(columns), 1)
    return Table(
        data, colWidths=[width] * len(columns), repeatRows=1, hAlign="LEFT",
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), ACCENT),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, WASH]),
            ("GRID", (0, 0), (-1, -1), 0.4, LINE),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]),
    )


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def _coerce(value: Any) -> Any:
    """Keep numbers numeric so Excel can actually sum and sort them."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    text = "" if value is None else str(value).strip()
    if not text:
        return ""
    compact = text.replace(",", "")
    if re.fullmatch(r"-?\d+", compact):
        try:
            return int(compact)
        except ValueError:
            return text
    if re.fullmatch(r"-?\d*\.\d+", compact):
        try:
            return float(compact)
        except ValueError:
            return text
    return text


def build_excel(payload: dict[str, Any], *, brand: str = "") -> tuple[Path, str]:
    """Render the sheet JSON to a formatted .xlsx. Returns (path, filename)."""
    ASSET_DIR.mkdir(parents=True, exist_ok=True)

    title = _clean(payload.get("title")) or "Resource"
    sheets = payload.get("sheets") or []
    if not sheets and payload.get("columns"):
        sheets = [payload]  # model returned a single flat sheet

    workbook = Workbook()
    workbook.remove(workbook.active)

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    title_font = Font(name="Calibri", bold=True, size=14, color="111827")
    note_font = Font(name="Calibri", size=9, color="6B7280", italic=True)
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for index, sheet in enumerate(sheets, 1):
        if not isinstance(sheet, dict):
            continue
        name = _clean(sheet.get("name")) or f"Sheet{index}"
        # Excel forbids []:*?/\ and caps names at 31 chars.
        name = re.sub(r"[\[\]:*?/\\]", "-", name)[:31] or f"Sheet{index}"
        ws = workbook.create_sheet(name)

        columns = [_clean(c) for c in sheet.get("columns") or []]
        rows = sheet.get("rows") or []

        ws.cell(row=1, column=1, value=_clean(sheet.get("title")) or title).font = title_font
        header_row = 3

        for col, heading in enumerate(columns, 1):
            cell = ws.cell(row=header_row, column=col, value=heading)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[header_row].height = 22

        for r, row in enumerate(rows, header_row + 1):
            values = row if isinstance(row, list) else [row]
            padded = (list(values) + [""] * len(columns))[: len(columns)] if columns else values
            for c, value in enumerate(padded, 1):
                cell = ws.cell(row=r, column=c, value=_coerce(value))
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Width from the longest cell, clamped so one long note cannot blow it out.
        for c, heading in enumerate(columns, 1):
            longest = len(str(heading))
            for r in range(header_row + 1, header_row + 1 + len(rows)):
                longest = max(longest, len(str(ws.cell(row=r, column=c).value or "")))
            ws.column_dimensions[get_column_letter(c)].width = min(max(longest + 4, 12), 52)

        if columns:
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
            ws.auto_filter.ref = (
                f"A{header_row}:{get_column_letter(len(columns))}{header_row + len(rows)}"
            )

        note = _clean(sheet.get("notes"))
        footer_row = header_row + len(rows) + 2
        if note:
            ws.cell(row=footer_row, column=1, value=note).font = note_font
            footer_row += 1
        ws.cell(
            row=footer_row, column=1,
            value=f"{brand + ' · ' if brand else ''}Generated {_timestamp()}",
        ).font = note_font

    if not workbook.sheetnames:
        ws = workbook.create_sheet("Sheet1")
        ws.cell(row=1, column=1, value=title).font = title_font

    filename = f"{slugify(title)}-{datetime.now(config.tz):%Y%m%d}.xlsx"
    path = ASSET_DIR / filename
    workbook.save(path)
    log.info("Built Excel %s (%s sheet(s))", filename, len(workbook.sheetnames))
    return path, filename
